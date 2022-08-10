import openpyxl

fpath = r'C:\DataCrawlingProjects\StartCoding\Chapter04\파이썬 엑셀 다루기\참가자_data.xlsx'
#1)엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

#2)엑셀 시트생성
ws = wb['대한민국 축구 국가대표팀']

#3)데이터 수집하기
ws['A3'] = 1
ws['B3'] = '김승규'

#3)엑셀 저장하기
wb.save(fpath)