from openpyxl import Workbook

#1) 엑셀 만들기
wb = Workbook()

#2) 엑셀 워크시트 만들기
ws = wb.create_sheet('대한민국 축구 국가대표팀')

#3) 데이터 추가하기
ws['A1'] = '등 번호'
ws['B1'] = '성명'

ws['A2'] = 7
ws['B2'] = '손흥민'

#4) 엑셀 저장하기 r : \를 하나만 쓰도록 해줌
wb.save(r"C:\DataCrawlingProjects\StartCoding\Chapter04\파이썬 엑셀 다루기\참가자_data.xlsx")