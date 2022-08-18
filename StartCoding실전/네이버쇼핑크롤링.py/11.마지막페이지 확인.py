from sqlite3 import Row
from turtle import title
import pyautogui
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment

keyWord = pyautogui.prompt("검색어를 입력하시오>>>")
lastPage = int(pyautogui.prompt("몇 페이지까지 크롤링 할까요?"))

#엑셀 생성
wb = Workbook()
#엑셀 시트 생성
ws = wb.create_sheet(keyWord)
#열 너비 조절
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120

#행 번호
row = 1

pageNum = 1
for i in range(1, int(lastPage)*10, 10):
    print(f"{pageNum}페이지 크롤링 중입니다.=================")
    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyWord}&start={i}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    articles = soup.select("div.info_group") #뉴스기사 10개 추출
    for article in articles:
        links = article.select("a.info")    # a태그인데 info class인 애들
        if len(links) >= 2: #링크가 2개 이상이면
            url = links[1].attrs['href']    #두 번째 링크의 URL 추출
            response = requests.get(url, headers={'User-agent' : 'Mozila/5.0'})
            html = response.text
            soup_sub = BeautifulSoup(html, 'html.parser') #다른 soup에 html담기
            #만약 스포츠뉴스 기사면
            if "sports" in response.url:
                title = soup_sub.select_one("title")
                content = soup_sub.select_one("#newsEndContents")
                divs = content.select("div")
                for div in divs:
                    div.decompose()
                paragraphs = content.select("p")
                for p in paragraphs:
                    p.decompose()
            #만약 연예뉴스 기사면
            elif "entertain" in response.url:
                title = soup_sub.select_one(".end_tit")
                content = soup_sub.select_one("#articeBody")
            else:
                title = soup_sub.select_one(".media_end_head_headline")
                content = soup_sub.select_one("#dic_area")
            # print(url)
            # print(title)
            # print(content.text.strip())
            ws[f'A{row}'] = url
            ws[f'B{row}'] = title.text.strip()
            ws[f'C{row}'] = content.text.strip()
            #자동 줄 바꿈
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            row = row + 1
            time.sleep(0.3)
    #마지막 페이지 여부 확인
    isLastPage = soup.select_one("a.btn_next").attrs['aria-disabled']
    if(isLastPage) == 'true':
        print("마지막 페이지입니다.")
        break
    pageNum = pageNum + 1
wb.save(f"{keyWord}_result.xlsx")