from openpyxl import Workbook

#1) 엑셀 만들기
wb = Workbook()

#2) 엑셀 워크시트 만들기
ws = wb.create_sheet('StartCoding')

#3) 데이터 추가하기
ws['A1'] = '스타트코딩'

#4) 엑셀 저장하기 r : \를 하나만 쓰도록 해줌
wb.save("test.xlsx")